import glob
import string
import io
import os
import requests

from datetime import datetime
from xml.etree import ElementTree as ET
from dateutil.parser import ParserError
import re

import openpyxl
import xml.etree.ElementTree as ET

import pandas as pd
from datetime import datetime, timedelta


from dateutil import parser

import argparse
import json
import os
import pathlib
import re
import requests
import shutil
import sys
import time

from helpers.config import Config


import os
import shutil



def group_element(_uid, group, cell_value):
   # print("group" + group)
    group_name = group.split("/")
   # print("Group name: " + group_name[0] + group_name[1])
    element = _uid.find('.//' + group_name[0])
   # print(element)
    if (element == None):
        element = ET.SubElement(_uid, group_name[0])
    if (group_name[0] != group_name[1]):
        group_element = ET.SubElement(element, group_name[1])
        group_element.text = cell_value
    return _uid


#All questions in repeat group should have an equal number of responses, even if a response is blank. 
#if question 1 in repeating group responses are {a, b, c}, question 2 responses need to have same number so indices match {1,,3}
def initial_repeat(_uid, col_arr, cell_value):
    #repeat/group_name/question
    group_name = col_arr[1]
    responses = cell_value.split(',')
    print(responses)
   
    elements = _uid.findall('.//' + group_name)
    print(elements)

    #initial group element doesn't exist
    if len(elements) == 0: 
        for response in responses:
            element = ET.SubElement(_uid, group_name)
            subelement = ET.SubElement(element, col_arr[2])
            subelement.text = response
    else: 
        #if (len(elements) != len(responses)): 
         #   raise Exception("Repeat group transfer failed. They should be saved in order with commas seperating each repeat response. No response should have a comma within it.")
        for group_el, response in zip(elements, responses):
            subelement = ET.SubElement(group_el, col_arr[2])
            subelement.text = response

    return _uid




#TODO: ideally would combine the two methods
def general_xls_to_xml(excel_file_path, xml_file_path, submission_data):
    workbook = openpyxl.load_workbook(excel_file_path)


    #select first sheet
    sheet = workbook.worksheets[0]

    uid = submission_data["asset_uid"]
    formhubuuid = submission_data["formhub_uuid"]
    v = submission_data["version"]
    __version__ = submission_data["__version__"]

    root = ET.Element("root") # Create the root element for the XML tree
    results = ET.Element("results")
   
    headers = [cell.value for cell in sheet[1]]
    print(headers)
    num_results = 0

    NSMAP = {"xmlns:jr" :  'http://openrosa.org/javarosa',
         "xmlns:orx" : 'http://openrosa.org/xforms', 
         "id" : str(uid),
         "version" : str(v)}

    # Iterate through rows and columns to populate XML
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # create formhub element with nested uuid
        _uid = ET.Element(uid, NSMAP)
        fhub_el = ET.SubElement(_uid, "formhub") 
        uuid_el = ET.SubElement(fhub_el, "uuid") 
        uuid_el.text = formhubuuid

        formatted_uuid = "uuid:"

        # Iterate through cells in the row and create corresponding XML elements
        for col_num, cell_value in enumerate(row, start=1):
                col_name = headers[col_num-1]
        
                if (col_name == "_uuid"):
               #     if cell_value == None:
                #        formatted_uuid = formatted_uuid + generate_new_instance_id()[1]
                 #   else:
                        formatted_uuid = formatted_uuid + str(cell_value)
            
                #TODO HOW TO FIX SUBMITTED_BY
                
                group_arr = col_name.split('/')

                if len(group_arr) == 2:
                    #create new element for ranking question
                    _uid = group_element(_uid, str(col_name), str(cell_value))
                    continue

                  #repeat groups are saved like this:
                #repeat/testname/group_question_1_text
                if len(group_arr) == 3 and group_arr[0] == "repeat":
                    _uid = initial_repeat(_uid, group_arr, str(cell_value))
                    continue

                if not (col_name.startswith("_")): 
                    cell_element = ET.SubElement(_uid, col_name)

                    if (col_name == "end" or col_name == "start"):
                        if (cell_value != None):
                            cell_value = cell_value.isoformat()
                
                    cell_element.text = str(cell_value)

        #TODO without repeatgropus/extra sheets this thing clears all of it oops     
        repeat_elements =  repeat_groups(_uid, formatted_uuid, excel_file_path)
        if (repeat_elements != None):
            _uid = repeat_elements
      

        version = ET.Element("__version__")
        version.text = (__version__)
        _uid.append(version)

        """meta tag before this ends
          <meta>
                <instanceID>uuid:a0ea37ef-ac71-434b-93b6-1713ef4c367f</instanceID>
                <deprecatedID>
            </meta>
        }"""
        meta = ET.Element("meta")
       # if (formatted_uuid == "uuid:"):
        #    formatted_uuid = generate_new_instance_id()[1]
        instanceId = ET.SubElement(meta, "instanceID") 
        deprecatedId = ET.SubElement(meta, "deprecatedID")

        instanceId.text = formatted_uuid
        deprecatedId.text = formatted_uuid
        
        _uid.append(meta)

        results.append(_uid)
        
        num_results += 1
    

    count =  ET.SubElement(root, 'count')
    count.text = (str(num_results))

    next = ET.SubElement(root, 'next')

    #next.text = submission_data["next"] #TODO

    previous = ET.SubElement(root, 'previous')
    #previous.text = submission_data["previous"] #TODO

    root.append(results)

    tree = ET.ElementTree(root)

    tree.write(xml_file_path) #for testing purposes

    workbook.close()
    return root

def repeat_groups(submission_xml, uuid, file_path): 
        uuid = uuid[len("uuid:"):]
        workbook = openpyxl.load_workbook(file_path)
        # Get the sheet names
        sheet_names = workbook.sheetnames
       
        sheet_names = sheet_names[1:]
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]

            submission_uid_header = headers.index("_submission__uuid")
        
            # Iterate through rows and columns to populate XML
            for row in sheet.iter_rows(min_row=2, values_only=True):
                #get the submission id for that row
                submission_uid = str(row[submission_uid_header])
                if (submission_uid == uuid):
                     element = ET.SubElement(submission_xml, sheet_name)

                # Iterate through cells in the row and create corresponding XML elements
                for col_num, cell_value in enumerate(row, start=1):
                    col_name = headers[col_num-1]
                    group_arr = col_name.split('/')
                    if len(group_arr) == 2 and sheet_name in col_name:
                         #check if the submission id matches
                        if (submission_uid == uuid): 
                            group_element = ET.SubElement(element, group_arr[1])
                            group_element.text = cell_value
                            

                            #submission_xml = group_element(submission_xml, str(col_name), str(cell_value))
            
            return submission_xml 


"""
def rename_media_folder(uuid, rowNum):
    current_attachments_path = os.path.join(
        Config.ATTACHMENTS_DIR, Config().src['asset_uid'], str(rowNum)
    )

    new_attachments_path = os.path.join(
        Config.ATTACHMENTS_DIR, Config().src['asset_uid'], str(uuid)
    )

    try:
        # Move the folder to the new path
        shutil.move(current_attachments_path, new_attachments_path)

    except Exception as e: #TODO
        print(f"Error: {e}")

"""
if __name__ == "__main__":
    data = {
        'asset_uid': 'asset_uid dummy',
        'version': 'version dummy',
        '__version__': '_v_',
        'formhub_uuid': 'formhub uuid dummy',
    }

    general_xls_to_xml("./matrixquestions.xlsx", './output.xml', data)


 
  #  rename_media_folder("FAKEUUID", 3)

    #workbook = openpyxl.load_workbook(file_path)
        
    # Get the sheet names
    #sheet_names = workbook.sheetnames
    
    #loop through the sheets, you might be able to finish the full xml like normal
    #now go into the next sheet (sheet number two is going to be named after the group_name)
    #somehow need to pass in the xml here FOR A SINGLE SUBMISSION HERE
    #check the name of that sheet
    #once you open the sheet, loop through the columns and rows
    #so when you have a column, you need to split it, and create an element with that group name
    #look into the submissionuid
    #if the submissionuid is the same as instanceid 
    #add to the xml element, else dont
    #lowkey you could try to use the same 


    #make sure the name is the same as after you split it 


